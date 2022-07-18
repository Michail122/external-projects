import smtplib
import openpyxl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
def send_email(message):
    server = smtplib.SMTP('smtp.yandex.com', 465)
    server.starttls()
    try:
        server.login(work_column_login, work_column_password)
        server.sendmail(work_column_login, work_column_login, f"Subject: CLICK ME PLEASE!\n{message}")
        return "The message was sent successfully!"
    except Exception as _ex:
        return f"{_ex}\nCheck your login or password please!"

def main():
    path = "sender_base.xlsx"  # Какой файл адресов для рассылки читаем?
    workbook = openpyxl.load_workbook(path)  # Собственно - читаем сам файл
    sheets_list = workbook.sheetnames  # Получаем список всех листов в книге
    global data_from_row, sheet, column_count, random_column, mail_adress_recipient, column_login, column_password, work_column_login, work_column_password, mail_server  # Делаем глобальные переменные (уточнить)
    sheet = workbook[sheets_list[0]]  # Делаем активным самый первый лист в книге
    column_count = sheet.max_row
    column = "1" #номер строки с логопасом
    column_login = "A"
    column_password = "B"
    work_column_login = column_login + column
    work_column_login= sheet[work_column_login].value
    work_column_password = column_password + column
    work_column_password=sheet[work_column_password].value
    message = input("Type your message: ")
    print(send_email(message=message))

if __name__ =="__main__":
    main()